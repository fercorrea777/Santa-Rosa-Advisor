"""
Ingesta de los informes mensuales de CADAM/DNRA (matriculacion e
importacion) a la base local SQLite (../data/cadam.db).

Modo automatico (recomendado, sin argumentos):
    python ingest_cadam.py

Escanea CADAM-DATA/ (en la raiz de PROYECTO CLAUDE), detecta para cada
archivo .xls/.xlsx si es matriculacion o importacion y el periodo
(AAAA-MM) a partir del NOMBRE del archivo, e ingesta todo lo que
encuentre. Reingestar el mismo periodo pisa los datos previos de ese
periodo (permite recargar si CADAM publica una revision) sin duplicar.

Modo manual (un archivo puntual, con periodo forzado):
    python ingest_cadam.py "<ruta al archivo>" <periodo AAAA-MM>

Convenciones de nombre de archivo reconocidas (case-insensitive):
    - Matriculacion: contiene "matricul"
      ej. "Informe de matriculacion de automotores junio_2026.xls"
    - Importacion: contiene "estadisticas totales" o "importac"
      ej. "Cadam - Estadisticas totales jun_2026 autos_camiones.xlsx"
    - Mes/anio: un mes en español (nombre completo o abreviado) seguido
      de un anio de 4 digitos, separados por "_", " " o "-"
      ej. "jun_2026", "junio-2026", "jun 2026"

Matriculacion: Cuadro 1 (por tipo) y Cuadro 2 (livianos por marca), en
formato largo (anio/mes/categoria) para poder acumular multiples
meses/anios.

Importacion: Cuadro 2 (por tipo, acumulado con comparacion interanual
ya calculada por CADAM: Acum anio anterior / Acum anio actual) y
Cuadro 3 (vehiculos livianos por tipo y mes del anio en curso).
"""
import re
import sys
import sqlite3
from pathlib import Path
import xlrd
import openpyxl

SCRIPT_DIR = Path(__file__).resolve().parent
DB_PATH = SCRIPT_DIR.parent / "data" / "cadam.db"
DEFAULT_CADAM_DATA_DIR = SCRIPT_DIR.parent.parent / "CADAM-DATA"

SCHEMA = """
CREATE TABLE IF NOT EXISTS informes (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    periodo TEXT NOT NULL,             -- AAAA-MM del informe (mes de cierre)
    tipo TEXT NOT NULL,                -- 'matriculacion' | 'importacion'
    archivo TEXT NOT NULL,
    fecha_ingesta TEXT NOT NULL DEFAULT (datetime('now')),
    UNIQUE(periodo, tipo)
);

CREATE TABLE IF NOT EXISTS matriculacion_tipo (
    informe_periodo TEXT NOT NULL,
    anio INTEGER NOT NULL,
    mes INTEGER NOT NULL,
    tipo TEXT NOT NULL,
    unidades INTEGER NOT NULL,
    PRIMARY KEY (informe_periodo, anio, mes, tipo)
);

CREATE TABLE IF NOT EXISTS matriculacion_marca (
    informe_periodo TEXT NOT NULL,
    anio INTEGER NOT NULL,
    mes INTEGER NOT NULL,
    marca TEXT NOT NULL,
    unidades INTEGER NOT NULL,
    PRIMARY KEY (informe_periodo, anio, mes, marca)
);

CREATE TABLE IF NOT EXISTS importacion_tipo_acum (
    informe_periodo TEXT NOT NULL,
    tipo TEXT NOT NULL,
    anio_actual INTEGER NOT NULL,
    anio_anterior INTEGER NOT NULL,
    unidades_actual INTEGER NOT NULL,
    unidades_anterior INTEGER NOT NULL,
    PRIMARY KEY (informe_periodo, tipo)
);

CREATE TABLE IF NOT EXISTS importacion_tipo_mensual (
    informe_periodo TEXT NOT NULL,
    anio INTEGER NOT NULL,
    mes INTEGER NOT NULL,
    tipo TEXT NOT NULL,
    unidades INTEGER NOT NULL,
    PRIMARY KEY (informe_periodo, anio, mes, tipo)
);

CREATE TABLE IF NOT EXISTS matriculacion_combustible_mensual (
    informe_periodo TEXT NOT NULL,
    anio INTEGER NOT NULL,
    mes INTEGER NOT NULL,
    combustible TEXT NOT NULL,
    unidades INTEGER NOT NULL,
    PRIMARY KEY (informe_periodo, anio, mes, combustible)
);

CREATE TABLE IF NOT EXISTS importacion_combustible_mensual (
    informe_periodo TEXT NOT NULL,
    anio INTEGER NOT NULL,
    mes INTEGER NOT NULL,
    combustible TEXT NOT NULL,
    unidades INTEGER NOT NULL,
    PRIMARY KEY (informe_periodo, anio, mes, combustible)
);

CREATE TABLE IF NOT EXISTS importacion_marca_mensual (
    informe_periodo TEXT NOT NULL,
    anio INTEGER NOT NULL,
    mes INTEGER NOT NULL,
    marca TEXT NOT NULL,
    unidades INTEGER NOT NULL,
    PRIMARY KEY (informe_periodo, anio, mes, marca)
);

CREATE TABLE IF NOT EXISTS matriculacion_modelo_mensual (
    informe_periodo TEXT NOT NULL,
    anio INTEGER NOT NULL,
    mes INTEGER NOT NULL,
    marca TEXT NOT NULL,
    modelo TEXT NOT NULL,
    unidades INTEGER NOT NULL,
    PRIMARY KEY (informe_periodo, anio, mes, marca, modelo)
);

CREATE TABLE IF NOT EXISTS importacion_modelo_mensual (
    informe_periodo TEXT NOT NULL,
    anio INTEGER NOT NULL,
    mes INTEGER NOT NULL,
    marca TEXT NOT NULL,
    modelo TEXT NOT NULL,
    unidades INTEGER NOT NULL,
    PRIMARY KEY (informe_periodo, anio, mes, marca, modelo)
);
"""

MONTH_COLS_CUADRO1 = [(1, 2, "ene"), (4, 5, "feb"), (7, 8, "mar"),
                       (10, 11, "abr"), (13, 14, "may"), (16, 17, "jun")]

MESES_ES = {
    "ene": 1, "enero": 1, "feb": 2, "febrero": 2, "mar": 3, "marzo": 3,
    "abr": 4, "abril": 4, "may": 5, "mayo": 5, "jun": 6, "junio": 6,
    "jul": 7, "julio": 7, "ago": 8, "agosto": 8,
    "sep": 9, "sept": 9, "set": 9, "septiembre": 9, "setiembre": 9,
    "oct": 10, "octubre": 10, "nov": 11, "noviembre": 11,
    "dic": 12, "diciembre": 12,
}
MESES_PATRON = "|".join(sorted(MESES_ES.keys(), key=len, reverse=True))
RE_PERIODO = re.compile(rf"(?i)\b({MESES_PATRON})[_\-\s]+(\d{{4}})\b")


def clean(s):
    return " ".join(str(s).split())


# CADAM usa nombres distintos para el mismo combustible segun el cuadro
# (ej. "Gasoil" en matriculacion vs. "Diesel" en importacion) -- se
# normaliza a una sola etiqueta para poder comparar/graficar ambas fuentes
# juntas (CLAUDE.md sec. 8).
COMBUSTIBLE_NORM = {
    "gasoil": "Diesel", "diesel": "Diesel", "diésel": "Diesel",
    "nafta": "Nafta",
    "flex": "Flex",
    "hibrido": "Híbrido", "híbrido": "Híbrido",
    "electrico": "Eléctrico", "eléctrico": "Eléctrico",
    "no definido": "No definido",
}
MESES_LARGOS = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
    "julio": 7, "agosto": 8, "setiembre": 9, "septiembre": 9,
    "octubre": 10, "noviembre": 11, "diciembre": 12,
}


def normalizar_combustible(label: str) -> str:
    key = clean(label).strip().lower()
    return COMBUSTIBLE_NORM.get(key, clean(label).strip())


def xldate_year(wb, serial):
    y, m, d, *_ = xlrd.xldate_as_tuple(serial, wb.datemode)
    return y, m


def detectar_tipo_y_periodo(nombre_archivo: str):
    """Devuelve (tipo, 'AAAA-MM') a partir del nombre del archivo, o (None, None)
    si no se puede clasificar con confianza."""
    lower = nombre_archivo.lower()
    if "matricul" in lower:
        tipo = "matriculacion"
    elif "estadisticas totales" in lower or "importac" in lower:
        tipo = "importacion"
    else:
        tipo = None

    m = RE_PERIODO.search(nombre_archivo)
    if not m:
        return tipo, None
    mes_num = MESES_ES[m.group(1).lower()]
    anio = int(m.group(2))
    return tipo, f"{anio:04d}-{mes_num:02d}"


# ---------------- MATRICULACION ----------------

def extract_cuadro1_matriculacion(wb):
    """Cuadro 1: matriculacion por tipo de automotor, formato largo."""
    sh = wb.sheet_by_name("1")
    date_row = sh.row_values(2)
    rows = []
    for r in range(3, sh.nrows):
        row = sh.row_values(r)
        if not row or not row[0]:
            continue
        label = clean(row[0])
        if label.upper().startswith(("ELABORADO", "FUENTE", "SUB TOTAL", "TOTAL")):
            if not label.upper().startswith(("ELABORADO", "FUENTE")):
                continue
            else:
                break
        for c25, c26, _tag in MONTH_COLS_CUADRO1:
            y25, m25 = xldate_year(wb, date_row[c25])
            y26, m26 = xldate_year(wb, date_row[c26])
            v25 = row[c25] if isinstance(row[c25], float) else 0
            v26 = row[c26] if isinstance(row[c26], float) else 0
            rows.append((y25, m25, label, int(v25)))
            rows.append((y26, m26, label, int(v26)))
    return rows


def extract_cuadro2_matriculacion(wb):
    """Cuadro 2: matriculacion de livianos por marca, formato largo."""
    sh = wb.sheet_by_name("2")
    date_row = sh.row_values(2)
    rows = []
    for r in range(3, sh.nrows):
        row = sh.row_values(r)
        if not row or not row[0] or not str(row[0]).strip():
            continue
        marca = clean(row[0])
        if marca.upper() in ("TOTAL",):
            continue
        if marca.upper().startswith(("ELABORADO", "FUENTE")):
            break
        for i in range(6):
            base = 1 + i * 3
            y25, m25 = xldate_year(wb, date_row[base])
            y26, m26 = xldate_year(wb, date_row[base + 1])
            v25 = row[base] if isinstance(row[base], float) else 0
            v26 = row[base + 1] if isinstance(row[base + 1], float) else 0
            rows.append((y25, m25, marca, int(v25)))
            rows.append((y26, m26, marca, int(v26)))
    return rows


def extract_cuadro19_matriculacion_combustible(wb):
    """Cuadro 19: matriculacion por tipo de movilidad (combustible), por mes,
    SOLO anio en curso (CADAM no publica el detalle mensual del anio
    anterior en este cuadro)."""
    sh = wb.sheet_by_name("19")
    title = clean(str(sh.cell_value(0, 0) or ""))
    m = re.search(r"(\d{4})", title)
    anio = int(m.group(1)) if m else None
    if anio is None:
        return []

    rows = []
    for r in range(2, sh.nrows):
        row = sh.row_values(r)
        label = row[0]
        if not label:
            continue
        label = clean(str(label))
        if label.upper().startswith(("TOTAL", "ELABORADO", "FUENTE")):
            if label.upper().startswith(("ELABORADO", "FUENTE")):
                break
            continue
        combustible = normalizar_combustible(label)
        for mes_idx in range(6):  # Ene..Jun -> columnas B..G (indices 1..6)
            val = row[1 + mes_idx]
            if isinstance(val, float):
                rows.append((anio, mes_idx + 1, combustible, int(val)))
    return rows


def extract_cuadro17_matriculacion_modelo(wb):
    """Cuadro 17: matriculacion por tipo, marca y modelo, SOLO anio en curso.
    La hoja original solo llena 'Marca' en la primera fila de cada marca
    (celdas fusionadas al exportar) -- hay que arrastrar el ultimo valor no
    vacio hacia abajo. CADAM aclara que los nombres de modelo son los
    originales de la DNRA, sin corregir."""
    sh = wb.sheet_by_name("17")
    title = clean(str(sh.cell_value(0, 0) or ""))
    m = re.search(r"(\d{4})", title)
    anio = int(m.group(1)) if m else None
    if anio is None:
        return []

    # El mismo (marca, modelo) puede aparecer en mas de una fila (variantes
    # de escritura que clean() unifica, o el mismo modelo repartido entre
    # mas de un "Grupo"/tipo) -- se suman en vez de tratarlas como unicas.
    acumulado = {}
    marca_actual = None
    for r in range(2, sh.nrows):
        row = sh.row_values(r)
        marca_celda = clean(str(row[1])) if row[1] else ""
        if marca_celda:
            marca_actual = marca_celda
        modelo = clean(str(row[2])) if row[2] else ""
        if not modelo:
            continue
        if modelo.upper().startswith(("TOTAL", "FUENTE", "ELABORADO")):
            break
        if marca_actual is None:
            continue
        for mes_idx in range(6):  # Ene..Jun -> columnas D..I (indices 3..8)
            val = row[3 + mes_idx]
            if isinstance(val, float) and val > 0:
                clave = (mes_idx + 1, marca_actual, modelo)
                acumulado[clave] = acumulado.get(clave, 0) + int(val)
    return [(anio, mes, marca, modelo, u) for (mes, marca, modelo), u in acumulado.items()]


# ---------------- IMPORTACION ----------------

def extract_cuadro2_importacion(wb):
    """Cuadro 2: importacion por tipo, acumulado con comparacion interanual
    ya calculada por CADAM (Acum anio anterior / Acum anio actual).
    wb: openpyxl Workbook (el archivo de importacion es .xlsx)."""
    sh = wb["2"]
    anio_actual = sh.cell(2, 3).value.year   # header: mes actual (col C, fila 2)
    anio_anterior = sh.cell(2, 5).value.year  # header: mismo mes anio anterior (col E, fila 2)

    rows = []
    for r in range(4, sh.max_row + 1):
        label = sh.cell(r, 1).value
        if not label:
            continue
        label = clean(str(label))
        if label.upper().startswith("FUENTE"):
            break
        if label.upper().startswith("TOTAL"):
            continue
        acum_anterior = sh.cell(r, 8).value
        acum_actual = sh.cell(r, 9).value
        if not isinstance(acum_anterior, (int, float)) or not isinstance(acum_actual, (int, float)):
            continue
        rows.append((label, anio_actual, anio_anterior, int(acum_actual), int(acum_anterior)))
    return rows


def extract_cuadro3_importacion_mensual(wb):
    """Cuadro 3: importacion de vehiculos livianos por tipo y mes, anio en curso.
    wb: openpyxl Workbook (el archivo de importacion es .xlsx)."""
    sh = wb["3"]
    title = clean(str(sh.cell(1, 1).value or ""))
    m = re.search(r"(\d{4})", title)
    anio = int(m.group(1)) if m else None

    rows = []
    for r in range(3, sh.max_row + 1):
        row = [sh.cell(r, c).value for c in range(1, 14)]
        label = row[0]
        if not label:
            continue
        label = clean(str(label))
        if label.upper().startswith(("TOTAL", "CREC", "FUENTE")):
            if label.upper().startswith("FUENTE"):
                break
            continue
        if anio is None:
            continue
        for mes_idx in range(6):  # Ene..Jun -> columnas B..G (indices 1..6 de row)
            val = row[1 + mes_idx]
            if isinstance(val, (int, float)):
                rows.append((anio, mes_idx + 1, label, int(val)))
    return rows


def extract_cuadro5_importacion_marca(wb):
    """Cuadro 5: importacion de vehiculos por marca y mes, SOLO anio en curso
    (mismo patron que el Cuadro 3: CADAM no publica la serie mensual del anio
    anterior en estos cuadros de detalle). wb: openpyxl Workbook (.xlsx)."""
    sh = wb["5"]
    title = clean(str(sh.cell(1, 1).value or ""))
    m = re.search(r"(\d{4})", title)
    anio = int(m.group(1)) if m else None
    if anio is None:
        return []

    rows = []
    for r in range(3, sh.max_row + 1):
        label = sh.cell(r, 1).value
        if not label:
            continue
        label = clean(str(label))
        if label.upper().startswith(("TOTAL", "FUENTE")):
            if label.upper().startswith("FUENTE"):
                break
            continue
        for mes_idx in range(6):  # Ene..Jun -> columnas B..G (indices 2..7)
            val = sh.cell(r, 2 + mes_idx).value
            if isinstance(val, (int, float)):
                rows.append((anio, mes_idx + 1, label, int(val)))
    return rows


def extract_cuadro8_importacion_modelo(wb):
    """Cuadro 8: importacion de vehiculos por modelo, por mes (y valor CIF,
    que no se usa aca), SOLO anio en curso. A diferencia del Cuadro 17 de
    matriculacion, aca 'Marca' viene completo en cada fila (no requiere
    arrastrar el valor hacia abajo)."""
    sh = wb["8"]
    title = clean(str(sh.cell(2, 1).value or ""))
    m = re.search(r"(\d{4})", title)
    anio = int(m.group(1)) if m else None
    if anio is None:
        return []

    # Igual que en Cuadro 17: por las dudas, sumar en vez de asumir que
    # (marca, modelo) es unico por fila.
    acumulado = {}
    for r in range(4, sh.max_row + 1):
        marca = sh.cell(r, 1).value
        if not marca:
            continue
        marca = clean(str(marca))
        if marca.upper().startswith(("TOTAL", "FUENTE")):
            if marca.upper().startswith("FUENTE"):
                break
            continue
        modelo = sh.cell(r, 2).value
        modelo = clean(str(modelo)) if modelo else ""
        if not modelo:
            continue
        for mes_idx in range(6):  # Ene..Jun -> columnas D..I (indices 4..9)
            val = sh.cell(r, 4 + mes_idx).value
            if isinstance(val, (int, float)) and val > 0:
                clave = (mes_idx + 1, marca, modelo)
                acumulado[clave] = acumulado.get(clave, 0) + int(val)
    return [(anio, mes, marca, modelo, u) for (mes, marca, modelo), u in acumulado.items()]


def extract_cuadro10_importacion_combustible(wb):
    """Cuadro 10A: importacion por tipo de combustible, por mes, SOLO anio en
    curso (mismo cuadro trae despues, indentado, el detalle por tipo de
    vehiculo dentro de cada mes -- se toman solo las filas de TOTAL mensual,
    no el desglose por tipo). wb: openpyxl Workbook (.xlsx)."""
    sh = wb["10"]
    title = clean(str(sh.cell(1, 1).value or ""))
    m = re.search(r"(\d{4})", title)
    anio = int(m.group(1)) if m else None
    if anio is None:
        return []

    # fila 2: 'Meses/Tipo de vehiculos', 'Nafta', None, 'Diesel', None, 'Flex',
    # None, 'Hibrido', None, 'Electrico', None, 'Total', ... -> col impar (B,D,F,H,J)
    combustibles_cols = []
    for c in range(2, 12, 2):
        label = sh.cell(2, c).value
        if label and str(label).strip().upper() != "TOTAL":
            combustibles_cols.append((c, normalizar_combustible(str(label))))

    rows = []
    for r in range(4, sh.max_row + 1):
        label = sh.cell(r, 1).value
        if not label:
            continue
        label = clean(str(label))
        if label.upper().startswith("FUENTE"):
            break
        mes = MESES_LARGOS.get(label.lower())
        if mes is None or mes > 6:  # solo filas de TOTAL mensual, Ene-Jun con datos reales
            continue
        for col, combustible in combustibles_cols:
            val = sh.cell(r, col).value
            if isinstance(val, (int, float)):
                rows.append((anio, mes, combustible, int(val)))
    return rows


# ---------------- DB ----------------

def get_conn():
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    con = sqlite3.connect(DB_PATH)
    con.executescript(SCHEMA)
    return con


def ingest_matriculacion(con, archivo: Path, periodo: str):
    wb = xlrd.open_workbook(str(archivo))
    tipo_rows = extract_cuadro1_matriculacion(wb)
    marca_rows = extract_cuadro2_matriculacion(wb)
    combustible_rows = extract_cuadro19_matriculacion_combustible(wb)
    modelo_rows = extract_cuadro17_matriculacion_modelo(wb)

    con.execute(
        "INSERT INTO informes (periodo, tipo, archivo) VALUES (?, 'matriculacion', ?) "
        "ON CONFLICT(periodo, tipo) DO UPDATE SET archivo=excluded.archivo, "
        "fecha_ingesta=datetime('now')",
        (periodo, str(archivo)),
    )
    con.execute("DELETE FROM matriculacion_tipo WHERE informe_periodo = ?", (periodo,))
    con.execute("DELETE FROM matriculacion_marca WHERE informe_periodo = ?", (periodo,))
    con.execute("DELETE FROM matriculacion_combustible_mensual WHERE informe_periodo = ?", (periodo,))
    con.execute("DELETE FROM matriculacion_modelo_mensual WHERE informe_periodo = ?", (periodo,))
    con.executemany(
        "INSERT INTO matriculacion_tipo (informe_periodo, anio, mes, tipo, unidades) "
        "VALUES (?, ?, ?, ?, ?)",
        [(periodo, y, m, t, u) for (y, m, t, u) in tipo_rows],
    )
    con.executemany(
        "INSERT INTO matriculacion_marca (informe_periodo, anio, mes, marca, unidades) "
        "VALUES (?, ?, ?, ?, ?)",
        [(periodo, y, m, mk, u) for (y, m, mk, u) in marca_rows],
    )
    con.executemany(
        "INSERT INTO matriculacion_combustible_mensual "
        "(informe_periodo, anio, mes, combustible, unidades) VALUES (?, ?, ?, ?, ?)",
        [(periodo, y, m, c, u) for (y, m, c, u) in combustible_rows],
    )
    con.executemany(
        "INSERT INTO matriculacion_modelo_mensual "
        "(informe_periodo, anio, mes, marca, modelo, unidades) VALUES (?, ?, ?, ?, ?, ?)",
        [(periodo, y, m, mk, mo, u) for (y, m, mk, mo, u) in modelo_rows],
    )
    con.commit()
    n_tipo = con.execute(
        "SELECT COUNT(*) FROM matriculacion_tipo WHERE informe_periodo=?", (periodo,)
    ).fetchone()[0]
    n_marca = con.execute(
        "SELECT COUNT(*) FROM matriculacion_marca WHERE informe_periodo=?", (periodo,)
    ).fetchone()[0]
    n_comb = con.execute(
        "SELECT COUNT(*) FROM matriculacion_combustible_mensual WHERE informe_periodo=?", (periodo,)
    ).fetchone()[0]
    n_modelo = con.execute(
        "SELECT COUNT(*) FROM matriculacion_modelo_mensual WHERE informe_periodo=?", (periodo,)
    ).fetchone()[0]
    print(f"  [matriculacion] {periodo} <- '{archivo.name}': {n_tipo} filas tipo, "
          f"{n_marca} filas marca, {n_comb} filas combustible, {n_modelo} filas modelo")


def ingest_importacion(con, archivo: Path, periodo: str):
    wb = openpyxl.load_workbook(str(archivo), data_only=True)
    acum_rows = extract_cuadro2_importacion(wb)
    mensual_rows = extract_cuadro3_importacion_mensual(wb)
    marca_rows = extract_cuadro5_importacion_marca(wb)
    combustible_rows = extract_cuadro10_importacion_combustible(wb)
    modelo_rows = extract_cuadro8_importacion_modelo(wb)

    con.execute(
        "INSERT INTO informes (periodo, tipo, archivo) VALUES (?, 'importacion', ?) "
        "ON CONFLICT(periodo, tipo) DO UPDATE SET archivo=excluded.archivo, "
        "fecha_ingesta=datetime('now')",
        (periodo, str(archivo)),
    )
    con.execute("DELETE FROM importacion_tipo_acum WHERE informe_periodo = ?", (periodo,))
    con.execute("DELETE FROM importacion_tipo_mensual WHERE informe_periodo = ?", (periodo,))
    con.execute("DELETE FROM importacion_marca_mensual WHERE informe_periodo = ?", (periodo,))
    con.execute("DELETE FROM importacion_combustible_mensual WHERE informe_periodo = ?", (periodo,))
    con.execute("DELETE FROM importacion_modelo_mensual WHERE informe_periodo = ?", (periodo,))
    con.executemany(
        "INSERT INTO importacion_tipo_acum "
        "(informe_periodo, tipo, anio_actual, anio_anterior, unidades_actual, unidades_anterior) "
        "VALUES (?, ?, ?, ?, ?, ?)",
        [(periodo, t, ya, yp, ua, up) for (t, ya, yp, ua, up) in acum_rows],
    )
    con.executemany(
        "INSERT INTO importacion_tipo_mensual (informe_periodo, anio, mes, tipo, unidades) "
        "VALUES (?, ?, ?, ?, ?)",
        [(periodo, y, m, t, u) for (y, m, t, u) in mensual_rows],
    )
    con.executemany(
        "INSERT INTO importacion_marca_mensual (informe_periodo, anio, mes, marca, unidades) "
        "VALUES (?, ?, ?, ?, ?)",
        [(periodo, y, m, mk, u) for (y, m, mk, u) in marca_rows],
    )
    con.executemany(
        "INSERT INTO importacion_combustible_mensual "
        "(informe_periodo, anio, mes, combustible, unidades) VALUES (?, ?, ?, ?, ?)",
        [(periodo, y, m, c, u) for (y, m, c, u) in combustible_rows],
    )
    con.executemany(
        "INSERT INTO importacion_modelo_mensual "
        "(informe_periodo, anio, mes, marca, modelo, unidades) VALUES (?, ?, ?, ?, ?, ?)",
        [(periodo, y, m, mk, mo, u) for (y, m, mk, mo, u) in modelo_rows],
    )
    con.commit()
    n_acum = con.execute(
        "SELECT COUNT(*) FROM importacion_tipo_acum WHERE informe_periodo=?", (periodo,)
    ).fetchone()[0]
    n_mensual = con.execute(
        "SELECT COUNT(*) FROM importacion_tipo_mensual WHERE informe_periodo=?", (periodo,)
    ).fetchone()[0]
    n_marca = con.execute(
        "SELECT COUNT(*) FROM importacion_marca_mensual WHERE informe_periodo=?", (periodo,)
    ).fetchone()[0]
    n_comb = con.execute(
        "SELECT COUNT(*) FROM importacion_combustible_mensual WHERE informe_periodo=?", (periodo,)
    ).fetchone()[0]
    n_modelo = con.execute(
        "SELECT COUNT(*) FROM importacion_modelo_mensual WHERE informe_periodo=?", (periodo,)
    ).fetchone()[0]
    print(f"  [importacion]   {periodo} <- '{archivo.name}': {n_acum} filas tipo (acum), "
          f"{n_mensual} filas mensuales, {n_marca} filas marca, {n_comb} filas combustible, "
          f"{n_modelo} filas modelo")


def scan_and_ingest(carpeta: Path):
    if not carpeta.exists():
        print(f"No existe la carpeta {carpeta}")
        sys.exit(1)

    archivos = sorted(
        p for p in carpeta.iterdir()
        if p.suffix.lower() in (".xls", ".xlsx") and not p.name.startswith("~$")
    )
    if not archivos:
        print(f"No hay archivos .xls/.xlsx en {carpeta}")
        return

    con = get_conn()
    print(f"Escaneando {carpeta} ({len(archivos)} archivo(s))...")
    for archivo in archivos:
        tipo, periodo = detectar_tipo_y_periodo(archivo.name)
        if tipo is None or periodo is None:
            print(f"  [omitido] '{archivo.name}': no se pudo detectar "
                  f"{'el tipo (matriculacion/importacion)' if tipo is None else 'el periodo (mes/anio)'} "
                  f"por el nombre del archivo.")
            continue
        try:
            if tipo == "matriculacion":
                ingest_matriculacion(con, archivo, periodo)
            else:
                ingest_importacion(con, archivo, periodo)
        except Exception as e:
            print(f"  [ERROR] '{archivo.name}' ({tipo}, {periodo}): {e}")
    con.close()
    print(f"Listo -> {DB_PATH}")


def main():
    if len(sys.argv) == 1:
        scan_and_ingest(DEFAULT_CADAM_DATA_DIR)
        return

    if len(sys.argv) != 3:
        print(__doc__)
        sys.exit(1)

    archivo = Path(sys.argv[1])
    periodo = sys.argv[2]
    if not archivo.exists():
        print(f"No existe: {archivo}")
        sys.exit(1)

    tipo, _ = detectar_tipo_y_periodo(archivo.name)
    if tipo is None:
        print(f"No se pudo detectar si '{archivo.name}' es matriculacion o importacion "
              f"por el nombre. Renombralo para incluir 'matricul' o 'estadisticas totales'.")
        sys.exit(1)

    con = get_conn()
    if tipo == "matriculacion":
        ingest_matriculacion(con, archivo, periodo)
    else:
        ingest_importacion(con, archivo, periodo)
    con.close()


if __name__ == "__main__":
    main()
